import re
import openpyxl, os
from io import BytesIO
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, Fill
from http.server import BaseHTTPRequestHandler 
from tempfile import NamedTemporaryFile
from PIL import Image 
from io import StringIO
import cgi

class handler(BaseHTTPRequestHandler):

    def do_POST(self):

        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={'REQUEST_METHOD':'POST', 'CONTENT_TYPE':self.headers['Content-Type'],})

        f = form['file'].file.read()
        bytes_data = BytesIO(f)

        im = Image.open(bytes_data)  
        wb = openpyxl.Workbook() 
        sheet = wb.active 

        W = int(form.getvalue("cols"))

        width, height   = im.size
        imgScale        = W/width
        newSize         = (int(width*imgScale), int(height*imgScale))

        im=im.resize(newSize)
        cols, rows      = im.size 

        for i in range(1, rows):
            sheet.row_dimensions[i].height = 6
        for j in range(1, cols):
            column_letter = get_column_letter(j)
            sheet.column_dimensions[column_letter].width = 1

        rgb_im = im.convert('RGB')

        for i in range(1, rows):
            for j in range(1, cols):
                c = tuple(rgb_im.getpixel((j, i)))
                rgb2hex = lambda r,g,b: f"ff{r:02x}{g:02x}{b:02x}"
                c = rgb2hex(*c)
                sheet.cell(row = i, column = j).value = " "
                customFill = PatternFill(start_color=c, end_color=c, fill_type='solid')
                sheet.cell(row = i, column = j).fill = customFill


        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            output = tmp.read()


        self.send_response(200)
        self.send_header("Accept-Ranges", "bytes")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Disposition", "attachment; filename=\"export.xls\"")
        self.send_header("Content-type", "application/vnd.ms-excel")
        self.end_headers()
        self.wfile.write(output)
        return  
