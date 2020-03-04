import re
import openpyxl, os
from io import BytesIO
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, Fill
from http.server import BaseHTTPRequestHandler 
from tempfile import NamedTemporaryFile
from PIL import Image 
from io import StringIO
import cgi

def build(input, output, cols=50):

    im = Image.open(input)  
    wb = openpyxl.Workbook() 
    sheet = wb.active 

    width, height   = im.size
    imgScale        = cols/width
    newSize         = (int(width*imgScale), int(height*imgScale))

    im=im.resize(newSize)
    cols, rows      = im.size

    for i in range(1, rows):
        sheet.row_dimensions[i].height = 6
    for j in range(1, cols):
        column_letter = get_column_letter(j)
        sheet.column_dimensions[column_letter].width = 1

    rgb_im = im.convert('RGB')

    for i in range(1, rows):
        for j in range(1, cols):
            c = tuple(rgb_im.getpixel((j, i)))
            rgb2hex = lambda r,g,b: f"ff{r:02x}{g:02x}{b:02x}"
            c = rgb2hex(*c)
            sheet.cell(row = i, column = j).value = " "
            customFill = PatternFill(start_color=c, end_color=c, fill_type='solid')
            sheet.cell(row = i, column = j).fill = customFill

    wb.save(output) 

 