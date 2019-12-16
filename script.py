import cv2
import re
import openpyxl, os
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, Fill

filename = 'monalisa.jpg'

oriimg = cv2.imread(filename)

wb = openpyxl.Workbook() 
sheet = wb.active 

W = 100
height, width, depth = oriimg.shape
imgScale = W/width

newX,newY = oriimg.shape[1]*imgScale, oriimg.shape[0]*imgScale
newimg = cv2.resize(oriimg,(int(newX),int(newY)))

rows, cols, _ = newimg.shape

for i in range(1, rows):
    sheet.row_dimensions[i].height = 1
for j in range(1, cols):
    column_letter = get_column_letter(j)
    sheet.column_dimensions[column_letter].width = 1

for i in range(1, rows):
    
    for j in range(1, cols):
        c = newimg[i, j] 

        rgb2hex = lambda r,g,b: f"ff{r:02x}{g:02x}{b:02x}"

        c = rgb2hex(c[0], c[1], c[2])
 
        #sheet.cell(row = i, column = j).value 
        customFill = PatternFill(start_color=c, end_color=c, fill_type='solid')
        sheet.cell(row = i, column = j).fill = customFill

wb.save('output.xlsx') 


# PatternFill(bgColor='FFEE08', fill_type = 'solid')